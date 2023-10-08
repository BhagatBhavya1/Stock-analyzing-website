import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
# import 

def save_to_excel():
    # Load the data from the CSV file and filter to get 'filtered_df'
    df2 = pd.read_csv('Downloads/fo24JUL2023bhav.csv', parse_dates=['EXPIRY_DT'])
    filtered_df = df2[df2['INSTRUMENT'].str.contains('FUT')]

    wb = load_workbook('./Copy.xlsx')
    ws = wb['Data']

    last_row = ws.max_row + 1

    date_style = NamedStyle(name='datetime', number_format='DD-MM-YYYY')

    for index, row in filtered_df.iterrows():
        target_row = last_row
        ex_date = row['EXPIRY_DT']
        formatted_date = ex_date.strftime("%d-%b-%Y")

        ws.cell(row=target_row, column=3, value=row['INSTRUMENT'])
        ws.cell(row=target_row, column=4, value=row['SYMBOL'])
        ws.cell(row=target_row, column=5, value=formatted_date).style = date_style
        ws.cell(row=target_row, column=6, value=row['STRIKE_PR'])
        ws.cell(row=target_row, column=7, value=row['OPTION_TYP'])
        ws.cell(row=target_row, column=8, value=row['OPEN'])
        ws.cell(row=target_row, column=9, value=row['HIGH'])
        ws.cell(row=target_row, column=10, value=row['LOW'])
        ws.cell(row=target_row, column=11, value=row['CLOSE'])
        ws.cell(row=target_row, column=12, value=row['SETTLE_PR'])
        ws.cell(row=target_row, column=13, value=row['CONTRACTS'])
        ws.cell(row=target_row, column=14, value=row['VAL_INLAKH'])
        ws.cell(row=target_row, column=15, value=row['OPEN_INT'])
        ws.cell(row=target_row, column=16, value=row['CHG_IN_OI'])
        
        time_stamp = row['TIMESTAMP']
        ws.cell(row=target_row, column=17, value=time_stamp).style = date_style

        last_row += 1

    wb.save('./Copy.xlsx')

save_to_excel()
